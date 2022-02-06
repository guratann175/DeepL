#使用ライブラリ
from tracemalloc import start
import deepl_get
import deepl_txt
import openpyxl
import time


start_time = time.process_time()
#CPU処理時間計測開始

deepl_txt.input_txt()

book = openpyxl.load_workbook('English.xlsx')
#openpyxlでエクセルファイルの読み込み
sheet = book['Sheet1']
#エクセルのシートの指定

for i in range(2,100):
    if(sheet.cell(row = i, column = 2).value == None):
        break
#読み込むインデックスの範囲を数字を入れて指定する。
    strings = sheet.cell(row = i, column = 2).value
#読み込みたいシートのセルを指定
    output = deepl_get.deepl(strings)
#自作関数(deep_getファイル参照)
    sheet.cell(row = i, column = 3).value = output
#特定のセルへの書き込み

book.save('English.xlsx')
#シートの保存　これを行わないと反映されません

deepl_get.close()

end_time = time.process_time()
print("処理時間", end_time - start_time)





